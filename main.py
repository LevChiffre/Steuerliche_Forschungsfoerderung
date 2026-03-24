import re
import pandas as pd
import holidays
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from calendar import monthrange, month_name

AVERAGE_WEEKS_PER_MONTH = 4.33


class cProject:
    def __init__(self, name, budget):
        self.name = name
        self.budget = budget

class cWeeklyWorkingTime:
    def __init__(self, monday:float=8.5, tuesday:float=8.5, wednesday:float=8.5, thursday:float=8.5, friday:float=6.0):
        self.monday = monday
        self.tuesday = tuesday
        self.wednesday =wednesday
        self.thursday = thursday
        self.friday = friday
        self.weekly_working_time = monday + tuesday + wednesday + thursday + friday

class cEmployees:
    def __init__(self, lastName:str, name:str, salary_year:int, project:cProject, workload:int=100, weekly_working_time:cWeeklyWorkingTime=cWeeklyWorkingTime()):
        self.lastName = lastName
        self.name = name
        self.salary_hour = salary_year / (12 * weekly_working_time.weekly_working_time * AVERAGE_WEEKS_PER_MONTH)
        self.project = project
        self.workload = workload
        self.weekly_working_time = weekly_working_time


# ======================== HELFERFUNKTIONEN ========================

def parse_absences(file_path):
    """
    Parst die Fehltage aus einer Excel-Datei.
    Erwartet strukturierte Pivot-Tabelle mit Blockköpfen:
       nummer - Nachname, Vorname (weiterer Text optional)
    """
    df = pd.read_excel(file_path, header=None)
    absences = {}
    for emp in EMPLOYEES:
        emp_name = f"{emp.lastName}, {emp.name}"
        absences[emp_name] = {}

    current_employee = None
    current_month = None
    unknown_blocks = []
    found_employees = set()

    for idx, row in df.iterrows():
        cell_0 = str(row[0]).strip() if pd.notna(row[0]) else ""

        found_employee = None
        m = re.match(r"^\s*\d+\s*-\s*(.+)$", cell_0)
        if m:
            candidate = m.group(1).strip()
            # nur Kandidaten mit "Nachname, Vorname" behandeln
            if "," in candidate:
                for emp in EMPLOYEES:
                    if re.search(r"\b" + re.escape(emp.lastName.lower()) + r"\b", candidate.lower()) and re.search(r"\b" + re.escape(emp.name.lower()) + r"\b", candidate.lower()):
                        found_employee = f"{emp.lastName}, {emp.name}"
                        break
                if not found_employee:
                    unknown_blocks.append(candidate)
            # Wenn kein Komma vorhanden ist, dann kein Mitarbeiter-Header, weitergehen

        if found_employee:
            current_employee = found_employee
            current_month = None
            found_employees.add(found_employee)
            continue

        if current_employee is None:
            continue

        if cell_0 in ['Jan', 'Feb', 'Mär', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']:
            month_map = {'Jan':1, 'Feb':2, 'Mär':3, 'Apr':4, 'Mai':5, 'Jun':6, 'Jul':7, 'Aug':8, 'Sep':9, 'Okt':10, 'Nov':11, 'Dez':12}
            current_month = month_map[cell_0]
            continue

        if isinstance(cell_0, str) and cell_0.startswith('202'):
            try:
                date = pd.to_datetime(cell_0).date()
                for col in range(1, 5):
                    if pd.notna(row[col]) and row[col] != 0:
                        reason = df.iloc[3, col]  # Header aus Zeile 3
                        absences[current_employee][date] = reason
            except Exception:
                continue

    if unknown_blocks:
        uniq = sorted(set(unknown_blocks))
        print("[WARN] Fehlende Zuordnung für diese Fehltage-Block-Namen:")
        for name in uniq:
            print("  -", name)

    expected = {f"{emp.lastName}, {emp.name}" for emp in EMPLOYEES}
    missing = sorted(expected - found_employees)
    if missing:
        print("[WARN] Keine Fehltage-Bloecke in der Datei fuer:")
        for name in missing:
            print("  -", name)

    return absences


def generate_all_days(year):
    """
    Generiert alle Tage eines Jahres als Liste von date-Objekten.
    """
    days = []
    start_date = datetime(year, 1, 1)
    end_date = datetime(year, 12, 31)
    current_date = start_date
    while current_date <= end_date:
        days.append(current_date.date())
        current_date += timedelta(days=1)
    return days



# ======================== DOKUMENTATION & KONFIGURATION ========================
"""
STUNDENNACHWEIS-GENERATOR
========================

ÜBERBLICK:
Dieses Programm erstellt automatisch Stundennachweise pro Projekt in Excel.
- Pro Projekt wird eine separate Excel-Datei erstellt
- Budget-Tracking pro Projekt und Mitarbeiter
- Fehlzeiten (Urlaub, Krankheit, etc.) werden aus externer Datei gelesen
- Arbeitstage filtern automatisch Wochenenden und Feiertage aus

KONFIGURATION:
-----------

1) JAHR EINSTELLEN:
   Suche nach "YEAR = " und setze das gewünschte Jahr (z.B. 2024):
   
   YEAR = 2024

2) PROJEKTE ANLEGEN:
   Pro Projekt eine Zeile hinzufügen mit: cProject("Projektname", Budget)
   Beispiel:
   
   project_1 = cProject("Project A", 200000)     # 200.000€ Budget
   project_2 = cProject("Project B", 300000)     # 300.000€ Budget
   project_3 = cProject("Project C", 150000)     # Neues Projekt
   
   Am Ende der Datei diese Projekte dann in die Ausführungsschleife eintragen:
   
   write_project_excel(project_1)
   write_project_excel(project_2)
   write_project_excel(project_3)  # <-- neu hinzufügen

3) MITARBEITER HINZUFÜGEN:
   Pro Mitarbeiter eine Zeile mit: cEmployees(Nachname, Vorname, Jahresgehalt, Projekt, Workload[%], ArbeitsZeiten)
   - Workload: Prozentsatz der Auslastung (50, 60, 100, etc.) für dieses Projekt
   - ArbeitsZeiten: optional, Standard ist Mo-Do: 8.5h, Fr: 6h (= 40h/Woche)
   
   Beispiel:
   
   employee_1 = cEmployees("Mustermann", "Max", 60000, project_1, 50)
               # Name: Max Mustermann, Gehalt: 60.000€/Jahr, 50% auf Project A
   
   employee_3 = cEmployees("Müller", "Anna", 55000, project_1, 80,
                            cWeeklyWorkingTime(8,8,8,8,6))
               # Name: Anna Müller, 80% auf Project A, 38h/Woche
   
   Dann EMPLOYEES-Liste aktualisieren:
   
   EMPLOYEES = [employee_1, employee_2, employee_3]  # alle Mitarbeiter auflisten

4) FEHLTAGE-DATEI:
   Datei: "Fehltage DIM 23-24.xlsx"
   Format: Pivot-Tabelle mit:
   - Spalte A: Mitarbeiternamen (Format: "Nachname, Vorname")
   - Spalte B (UT): Urlaub
   - Spalte C (KT): Krankheit
   - Spalte D (GT): Sonstiges Urlaub
   - Spalte E (KiKr): Kinderkrankheit
   
   Pro Mitarbeiter-Block ein Monatszyklus mit Daten (1 oder 0).
   Programm liest automatisch alle Mitarbeiter aus dieser Datei.

STUNDENLOHN-BERECHNUNG:
Stundenlohn wird automatisch aus Jahresgehalt berechnet:
   Stundenlohn = Jahresgehalt / (12 Monate × Wochenstunden × 4.33 Wochen/Monat)

AUSGABE:
Pro Projekt eine Excel-Datei: "Stundennachweis_Projektname_2024.xlsx"
   - Spalten: Mo-Fr + fixe Gesamtspalten
   - Zeilen: Mitarbeiter pro Monat
   - Farben: Rot = Fehlzeiten, Hellrot = Budget überschritten
   - Budgetkumulativ: Läuft über alle Monate (setzt nicht zurück)
"""


# ======================== GLOBALE EINSTELLUNGEN ========================

MONTHS_GER = ["dummy","Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
DAY_SHORT_GER = ["Mo", "Di", "Mi", "Do", "Fr"]

# ==================== JAHR & DATEINAME ====================
YEAR = 2024  # <-- HIER DAS JAHR EINSTELLEN
# =========================================================

EXCEL_NAME = "Stundennachweis_" + str(YEAR) + ".xlsx"
FEHLTAGE_FILE = "Fehltage DIM 23-24.xlsx"

# Feiertage für Hessen laden
hessian_holidays = holidays.DE(prov='HE', years=YEAR)

# ==================== PROJEKTE DEFINIEREN ====================
project_1 = cProject("Project A", 20000)
project_2 = cProject("Project B", 30000)
# Weitere Projekte hier hinzufügen...
# ============================================================

# ==================== MITARBEITER DEFINIEREN ====================
employee_1 = cEmployees("Mustermann", "Max", 60000, project_1, 50, cWeeklyWorkingTime())
employee_2 = cEmployees("Zufall", "Reiner", 50000, project_1, 100, cWeeklyWorkingTime())
employee_3 = cEmployees("Maduschen", "Isolde", 150000, project_2, 50, cWeeklyWorkingTime(6,4,8,6,4))
employee_4 = cEmployees("Silie", "Peter", 50000, project_2, 100, cWeeklyWorkingTime())
employee_5 = cEmployees("Eder", "Meister", 70000, project_2, 50, cWeeklyWorkingTime())

# Weitere Mitarbeiter hier hinzufügen...
EMPLOYEES = [employee_1, employee_2, employee_3, employee_4, employee_5]
# ================================================================

# Fehltage laden
absences = parse_absences(FEHLTAGE_FILE)


# ======================== HAUPTFUNKTION ========================

def write_project_excel(project):
    employees = [emp for emp in EMPLOYEES if emp.project == project]
    if not employees:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = project.name

    # Jahr-Überschrift
    ws['A1'] = f"Projekt: {project.name} ({YEAR})"
    ws['A1'].font = Font(bold=True, size=14)

    # Stile
    red_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    holiday_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    current_row = 3
    total_project_cost = 0.0
    project_budget_cumulative = 0.0

    for month in range(1, 13):
        month_name_str = MONTHS_GER[month]
        ws[f'A{current_row}'] = month_name_str
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = alignment_center
        current_row += 1

        # Arbeitstage sammeln
        workdays = []
        num_days = monthrange(YEAR, month)[1]
        for day in range(1, num_days + 1):
            date = datetime(YEAR, month, day).date()
            if date.weekday() >= 5:
                continue
            weekday = DAY_SHORT_GER[date.weekday()]
            is_holiday = date in hessian_holidays
            workdays.append({'day': day, 'weekday': weekday, 'date': date, 'is_holiday': is_holiday})

        # Kopfzeilen für Tage
        ws[f'A{current_row}'] = "Tag"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = alignment_left
        for idx, d in enumerate(workdays):
            cell = ws.cell(row=current_row, column=2 + idx, value=d['day'])
            cell.font = header_font
            cell.alignment = alignment_center
            if d.get('is_holiday'):
                cell.fill = holiday_fill
        # fixe Gesamtspalten unabhängig von workdays-Length
        total_col = 2 + 31
        ws.cell(row=current_row, column=total_col, value="Summe Std.").font = header_font
        ws.cell(row=current_row, column=total_col).alignment = alignment_center
        ws.cell(row=current_row, column=total_col+1, value="Stundenlohn").font = header_font
        ws.cell(row=current_row, column=total_col+1).alignment = alignment_center
        # Spalte total_col+2 bleibt leer (Abstand)
        ws.cell(row=current_row, column=total_col+3, value="Summe x Stundenlohn").font = header_font
        ws.cell(row=current_row, column=total_col+3).alignment = alignment_center
        ws.cell(row=current_row, column=total_col+4, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+5, value="Aufgebrauchtes Budget").font = header_font
        ws.cell(row=current_row, column=total_col+5).alignment = alignment_center
        current_row += 1

        ws[f'A{current_row}'] = "Wochentag"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = alignment_left
        for idx, d in enumerate(workdays):
            cell = ws.cell(row=current_row, column=2 + idx, value=d['weekday'])
            cell.alignment = alignment_center
            if d.get('is_holiday'):
                cell.fill = holiday_fill
        # Gesamtspalten bleiben ebenfalls fix
        ws.cell(row=current_row, column=total_col, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+1, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+2, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+3, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+4, value="").alignment = alignment_center
        ws.cell(row=current_row, column=total_col+5, value="").alignment = alignment_center
        current_row += 1

        # Projekttageskosten und Zeilenpositionen vorbereiten
        start_employee_row = current_row
        for emp in employees:
            current_row += 1

        project_day_costs = {d['date']: 0.0 for d in workdays}

        # 1. Kosten kalkulieren: für jeden Mitarbeiter Tageswert (exclude Abwesenheit)
        emp_day_values = {}
        for emp in employees:
            emp_name = f"{emp.lastName}, {emp.name}"
            emp_day_values[emp_name] = {}
            for d in workdays:
                date = d['date']
                weekday_idx = date.weekday()
                day_hours_map = [
                    emp.weekly_working_time.monday,
                    emp.weekly_working_time.tuesday,
                    emp.weekly_working_time.wednesday,
                    emp.weekly_working_time.thursday,
                    emp.weekly_working_time.friday,
                ]
                project_hours = day_hours_map[weekday_idx] * (emp.workload / 100.0)
                project_hours = round(project_hours * 2) / 2
                project_hours = round(project_hours, 1)
                day_cost = project_hours * emp.salary_hour

                is_absence = date in absences.get(emp_name, {})
                if is_absence:
                    reason = absences[emp_name][date]
                elif d.get('is_holiday'):
                    # Feiertag: als Urlaub behandeln
                    reason = "UT"
                else:
                    reason = None

                if reason is not None:
                    # Fehlzeiten gelten budgetwirksam als Stunden, aber optisch markiert
                    emp_day_values[emp_name][date] = {'hours': project_hours, 'reason': reason, 'cost': day_cost}
                else:
                    emp_day_values[emp_name][date] = {'hours': project_hours, 'reason': None, 'cost': day_cost}

                # Auch Feiertage (UT) sollen in Stunden/Kosten eingehen
                project_day_costs[date] += day_cost

        # 2. Budgetcheck pro Tag, Marke einsetzen BUDGET! wenn überschritten insgesamt
        over_budget_days = set()
        budget_exhausted = False
        for date in sorted(project_day_costs.keys()):
            day_cost = project_day_costs[date]
            if budget_exhausted or (total_project_cost + day_cost > project.budget):
                over_budget_days.add(date)
                budget_exhausted = True
            else:
                total_project_cost += day_cost

        # 3. Schreibe Personen und Zellen
        current_row = start_employee_row
        for emp in employees:
            emp_name = f"{emp.lastName}, {emp.name}"
            ws[f'A{current_row}'] = emp_name
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].alignment = alignment_left

            row_hours = 0.0
            row_cost = 0.0
            for idx, d in enumerate(workdays):
                date = d['date']
                col = 2 + idx
                info = emp_day_values[emp_name][date]

                if date in over_budget_days:
                    cell = ws.cell(row=current_row, column=col, value='')
                elif info['reason'] is not None:
                    cell = ws.cell(row=current_row, column=col, value=info['reason'])
                    cell.fill = red_fill
                    # Fehlzeiten sollen in der Summe/Budget mitgezaehlt werden
                    row_hours += info['hours'] if info['hours'] is not None else 0
                    row_cost += info['cost']
                elif d.get('is_holiday'):
                    cell = ws.cell(row=current_row, column=col, value="")
                else:
                    cell = ws.cell(row=current_row, column=col, value=float(info['hours']))
                    row_hours += info['hours'] if info['hours'] is not None else 0
                    row_cost += info['cost']

                cell.alignment = alignment_center
                # Immer eine Nachkommastelle anzeigen, damit 0.5 sichtbar ist
                cell.number_format = "0.0"

            # Zusatzspalten pro Mitarbeiter (fixierte Gesamtspalten bei 31 Tage)
            total_col = 2 + 31
            cell_sum = ws.cell(row=current_row, column=total_col, value=round(row_hours, 1))
            cell_sum.alignment = alignment_center
            cell_sum.number_format = "0.0"

            # Stundenlohn (Spalte 34)
            cell_wage = ws.cell(row=current_row, column=total_col+1, value=round(emp.salary_hour, 2))
            cell_wage.alignment = alignment_center

            # Spalte 35 bleibt leer (Abstand)

            # Summe x Stundenlohn (Spalte 36): Produkt aus Summe Std. und Stundenlohn
            row_cost = row_hours * emp.salary_hour
            cell_budget = ws.cell(row=current_row, column=total_col+3, value=round(row_cost, 2))
            cell_budget.alignment = alignment_center

            # Abstandsspalte (Spalte 37) bleibt leer

            # Aufgebrauchtes Budget: kumulierte Gesamtsumme aller "Summe x Stundenlohn"
            project_budget_cumulative += row_cost
            cell_project_budget = ws.cell(row=current_row, column=total_col+5, value=round(project_budget_cumulative, 2))
            cell_project_budget.alignment = alignment_center

            current_row += 1

        current_row += 2

        current_row += 2

    # Spaltenbreite
    ws.column_dimensions['A'].width = 30
    for c in range(2, 50):
        ws.column_dimensions[get_column_letter(c)].width = 5

    # Breitere Spalten für Zusammenfassung
    total_col = 2 + 31
    ws.column_dimensions[get_column_letter(total_col)].width = 14      # Summe Std.
    ws.column_dimensions[get_column_letter(total_col+1)].width = 14    # Stundenlohn
    ws.column_dimensions[get_column_letter(total_col+3)].width = 20    # Summe x Stundenlohn
    ws.column_dimensions[get_column_letter(total_col+5)].width = 20    # Aufgebrauchtes Budget

    output_file = f"Stundennachweis_{project.name}_{YEAR}.xlsx"
    try:
        wb.save(output_file)
        print(f"Projektdatei {output_file} erstellt. Gesamtkosten: {total_project_cost:.2f} / Budget: {project.budget}")
    except Exception as e:
        print(f"Fehler beim Speichern von {output_file}: {e}")


# ======================== AUSFÜHRUNG ========================
write_project_excel(project_1)
write_project_excel(project_2)
# Weitere Projekte hier hinzufügen...
